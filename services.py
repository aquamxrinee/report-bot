import re
import shutil
import logging
import requests
import pandas as pd
import openpyxl
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from apscheduler.schedulers.background import BackgroundScheduler
import asyncio
import os

from config import TEMP_DIR, logger, DB_PATH, ALLOWED_USERS
from models import save_report_to_db, get_report_id_by_period, calculate_file_hash

scheduler = BackgroundScheduler()

def detect_report_type(filename):
    name = filename.lower()
    if 'осн' in name or 'osn' in name:
        return 'osn'
    elif 'вык' in name or 'vyk' in name:
        return 'vyk'
    return None

def parse_date_from_period(date_period):
    try:
        parts = date_period.split('-')
        start = parts[0].strip()
        end = parts[1].strip()
        year = datetime.now().year
        start_dt = datetime.strptime(start + f".{year}", "%d.%m.%Y")
        end_dt = datetime.strptime(end + f".{year}", "%d.%m.%Y")
        return start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")
    except:
        return None, None

class ReportProcessor:
    def process_files(self, osn_path, vyk_path, template_path):
        df_osn = pd.read_excel(osn_path)
        df_vyk = pd.read_excel(vyk_path)
        logger.info(f"Колонки основного: {df_osn.columns.tolist()}")
        logger.info(f"Колонки выкупов: {df_vyk.columns.tolist()}")
        filename = Path(osn_path).name
        match = re.search(r'(\d{1,2})\.(\d{2})-(\d{1,2})\.(\d{2})', filename)
        date_range = f"{match.group(1)}.{match.group(2)}-{match.group(3)}.{match.group(4)}" if match else datetime.now().strftime("%d.%m")
        values = self._calculate_all_values(df_osn, df_vyk, date_range)
        self._fill_template(template_path, values)
        articles = self._get_articles_stats(df_osn, df_vyk)
        return values, articles, date_range

    def _get_articles_stats(self, df_osn, df_vyk):
        result = {}
        def normalize_cols(df):
            return {str(col).strip().lower(): col for col in df.columns}
        cols_osn = normalize_cols(df_osn)
        cols_vyk = normalize_cols(df_vyk)
        all_cols = {**cols_vyk, **cols_osn}
        qty_variants = ['количество', 'кол-во', 'количество товара', 'кол-во (шт.)', 'кол-во шт', 'quantity', 'количество,шт']
        art_variants = [
            'артикул поставщика', 'артикул', 'артикул товара', 'номенклатура',
            'sku', 'артикул(поставщика)', 'артикул поставщика (поставщика)',
            'код товара', 'id товара', 'vendor code', 'article'
        ]
        nm_id_variants = ['код номенклатуры', 'nmId', 'nm_id', 'артикул товара', 'номенклатура', 'артикул']
        qty_col = art_col = nm_id_col = None
        for v in qty_variants:
            if v in all_cols:
                qty_col = all_cols[v]
                break
        for v in art_variants:
            if v in all_cols:
                art_col = all_cols[v]
                break
        for v in nm_id_variants:
            if v in all_cols:
                nm_id_col = all_cols[v]
                break
        if art_col is None:
            original_cols = df_osn.columns.tolist()
            for col in original_cols:
                col_lower = col.strip().lower()
                if any(keyword in col_lower for keyword in ['артикул поставщика', 'vendor', 'article', 'артикул']):
                    if 'номенклатура' not in col_lower and 'код' not in col_lower:
                        art_col = col
                        logger.info(f"🔍 Нашли колонку артикула по точному совпадению: '{col}'")
                        break
        if nm_id_col is None:
            original_cols = df_osn.columns.tolist()
            for col in original_cols:
                if 'Код номенклатуры' in col or col.strip().lower() == 'код номенклатуры':
                    nm_id_col = col
                    logger.info(f"🔍 Нашли колонку nm_id по точному совпадению: '{col}'")
                    break
        if qty_col is None:
            logger.warning(f"❌ Колонка количества не найдена. Доступные: {list(all_cols.keys())}")
            return result
        if art_col is None:
            logger.warning(f"❌ Колонка артикула не найдена. Доступные: {list(all_cols.keys())}")
            possible = [col for col in all_cols.keys() if 'артикул' in col or 'article' in col]
            if possible:
                art_col = all_cols[possible[0]]
                logger.info(f"✅ Нашли возможную колонку артикула: '{art_col}'")
        logger.info(f"✅ Найдены колонки: количество='{qty_col}', артикул='{art_col}', nm_id='{nm_id_col}'")
        for df, key in [(df_osn, 'sales'), (df_vyk, 'vyk')]:
            for bren, mask_func in [
                ('Цап царапкин', lambda d: (d['Бренд'] == 'Цап царапкин') | (d['Бренд'].isna())),
                ('Harakiri', lambda d: d['Бренд'] == 'Harakiri')
            ]:
                mask = mask_func(df)
                df_bren = df[mask]
                if df_bren.empty:
                    continue
                sales = df_bren[(df_bren['Тип документа'] == 'Продажа') & (df_bren[qty_col] > 0)]
                if sales.empty:
                    articles = {}
                else:
                    agg_dict = {'quantity': (qty_col, 'sum'), 'revenue': ('Цена розничная', 'sum')}
                    if nm_id_col:
                        agg_dict['nm_id'] = (nm_id_col, 'first')
                    agg_sales = sales.groupby(art_col).agg(**agg_dict).to_dict('index')
                    articles = {}
                    for art, vals in agg_sales.items():
                        nm_id_val = vals.get('nm_id') if nm_id_col else None
                        if nm_id_val is not None:
                            try:
                                nm_id_val = int(float(nm_id_val))
                            except:
                                nm_id_val = None
                        articles[art] = {'quantity': vals['quantity'], 'revenue': vals['revenue'], 'nm_id': nm_id_val}
                if bren not in result:
                    result[bren] = {}
                result[bren][key] = articles
        logger.info(f"📦 Собрано артикулов: {sum(len(v.get('sales', {})) for v in result.values())}")
        return result

    def _calculate_all_values(self, df_osn, df_vyk, date_range):
        values = {'B1': date_range, 'F1': date_range}
        mask_carp_sale = ((df_osn['Бренд'] == 'Цап царапкин') | (df_osn['Бренд'].isna())) & (df_osn['Тип документа'] == 'Продажа')
        values['B4'] = df_osn[mask_carp_sale]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_carp_return = ((df_osn['Бренд'] == 'Цап царапкин') | (df_osn['Бренд'].isna())) & (df_osn['Тип документа'] == 'Возврат')
        values['B5'] = df_osn[mask_carp_return]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_carp_all = (df_osn['Бренд'] == 'Цап царапкин') | (df_osn['Бренд'].isna())
        values['B7'] = df_osn[mask_carp_all]['Услуги по доставке товара покупателю'].sum()
        values['B9'] = df_osn[mask_carp_all]['Операции на приемке'].sum()
        values['B10'] = df_osn['Общая сумма штрафов'].sum()
        values['B11'] = df_osn[mask_carp_all]['Удержания'].sum()
        values['B26'] = df_osn[mask_carp_all]['Хранение'].sum()
        values['B29'] = df_osn[mask_carp_all]['Разовое изменение срока перечисления денежных средств'].sum()
        values['B44'] = df_osn[mask_carp_sale]['Цена розничная'].sum()
        mask_hara_sale = (df_osn['Бренд'] == 'Harakiri') & (df_osn['Тип документа'] == 'Продажа')
        values['F4'] = df_osn[mask_hara_sale]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_hara_return = (df_osn['Бренд'] == 'Harakiri') & (df_osn['Тип документа'] == 'Возврат')
        values['F5'] = df_osn[mask_hara_return]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_hara_all = df_osn['Бренд'] == 'Harakiri'
        values['F7'] = df_osn[mask_hara_all]['Услуги по доставке товара покупателю'].sum()
        values['F9'] = df_osn[mask_hara_all]['Операции на приемке'].sum()
        values['F10'] = df_osn[mask_hara_all]['Общая сумма штрафов'].sum()
        values['F11'] = df_osn[mask_hara_all]['Удержания'].sum()
        values['B32'] = df_osn[mask_hara_sale]['Цена розничная'].sum()
        mask_carp_vyk_sale = ((df_vyk['Бренд'] == 'Цап царапкин') | (df_vyk['Бренд'].isna())) & (df_vyk['Тип документа'] == 'Продажа')
        values['M4'] = df_vyk[mask_carp_vyk_sale]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_carp_vyk_return = ((df_vyk['Бренд'] == 'Цап царапкин') | (df_vyk['Бренд'].isna())) & (df_vyk['Тип документа'] == 'Возврат')
        values['M5'] = df_vyk[mask_carp_vyk_return]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_carp_vyk_all = (df_vyk['Бренд'] == 'Цап царапкин') | (df_vyk['Бренд'].isna())
        values['M7'] = df_vyk[mask_carp_vyk_all]['Услуги по доставке товара покупателю'].sum()
        values['M8'] = df_vyk[mask_carp_vyk_all]['Операции на приемке'].sum()
        values['M9'] = df_vyk['Общая сумма штрафов'].sum()
        values['B47'] = df_vyk[mask_carp_vyk_sale]['Цена розничная'].sum()
        mask_hara_vyk_sale = (df_vyk['Бренд'] == 'Harakiri') & (df_vyk['Тип документа'] == 'Продажа')
        values['Q4'] = df_vyk[mask_hara_vyk_sale]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_hara_vyk_return = (df_vyk['Бренд'] == 'Harakiri') & (df_vyk['Тип документа'] == 'Возврат')
        values['Q5'] = df_vyk[mask_hara_vyk_return]['К перечислению Продавцу за реализованный Товар'].sum()
        mask_hara_vyk_all = df_vyk['Бренд'] == 'Harakiri'
        values['Q7'] = df_vyk[mask_hara_vyk_all]['Услуги по доставке товара покупателю'].sum()
        values['Q8'] = df_vyk[mask_hara_vyk_all]['Операции на приемке'].sum()
        values['Q9'] = df_vyk[mask_hara_vyk_all]['Общая сумма штрафов'].sum()
        values['B41'] = df_vyk[mask_hara_vyk_sale]['Цена розничная'].sum()
        col = "Размер компенсации платёжных услуг/Комиссии за интеграцию платёжных сервисов, %"
        if col in df_osn.columns:
            filtered = df_osn[col][df_osn[col].notna() & (df_osn[col] > 0)]
            if not filtered.empty:
                values['B56'] = filtered.mean()
                values['B59'] = filtered.median()
                values['B62'] = filtered.min()
                values['B65'] = filtered.max()
            else:
                values['B56'] = values['B59'] = values['B62'] = values['B65'] = 0
        else:
            values['B56'] = values['B59'] = values['B62'] = values['B65'] = 0
        return values

    def _fill_template(self, template_path, values):
        wb = openpyxl.load_workbook(template_path, data_only=False, keep_links=False, keep_vba=False)
        ws = wb.active
        for cell, value in values.items():
            ws[cell] = value
            if isinstance(value, float) and value != int(value):
                ws[cell].number_format = '0.00'
        ws.sheet_view.calcMode = 'manual'
        wb.save(template_path)


def extract_metrics_from_values(values):
    return {
        'avg_acquiring': values.get('B56', 0),
        'wb_carp': values.get('B44', 0),
        'wb_hara': values.get('B32', 0),
        'wb_total': values.get('B44', 0) + values.get('B32', 0),
        'k_vyvodu_carp': values.get('B4', 0),
        'k_vyvodu_hara': values.get('F4', 0),
    }

def prepare_api_dataframe(detail_list):
    df = pd.DataFrame(detail_list)
    numeric_cols = ['retailAmount', 'forPay', 'quantity', 'penalty', 'deliveryAmount',
                    'paidAcceptance', 'paidStorage', 'deduction', 'acquiringPercent',
                    'additionalPayment']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            df[col] = 0
    df['Бренд'] = df.get('brandName', '')
    df['Тип документа'] = df.get('docTypeName', '')
    df['Цена розничная'] = df.get('retailAmount', 0)
    df['К перечислению Продавцу за реализованный Товар'] = df.get('forPay', 0)
    df['Общая сумма штрафов'] = df.get('penalty', 0)
    df['Услуги по доставке товара покупателю'] = df.get('deliveryAmount', 0)
    df['Операции на приемке'] = df.get('paidAcceptance', 0)
    df['Хранение'] = df.get('paidStorage', 0)
    df['Удержания'] = df.get('deduction', 0)
    df['Разовое изменение срока перечисления денежных средств'] = df.get('additionalPayment', 0)
    df['Количество'] = df.get('quantity', 0).astype(int)
    df['acquiring_percent'] = df.get('acquiringPercent', 0)
    df['Размер компенсации платёжных услуг/Комиссии за интеграцию платёжных сервисов, %'] = df['acquiring_percent']
    return df


async def process_auto_report(app, osn_detail, vyk_detail, period_str, date_from, date_to):
    all_detail = osn_detail + vyk_detail
    df = prepare_api_dataframe(all_detail)
    processor = ReportProcessor()
    values = processor._calculate_all_values(df, df, period_str)
    template_path = Path("шаблон.xlsx")
    processor._fill_template(template_path, values)
    file_hash = hashlib.md5(f"{date_from}_{date_to}".encode()).hexdigest()

    from wb_api import get_buyout_by_brands
    buyouts = {}
    try:
        buyouts = get_buyout_by_brands(date_from, date_to)
    except Exception as e:
        logger.error(f"Ошибка получения выкупов в автоотчёте: {e}")

    metrics = extract_metrics_from_values(values)
    metrics['buyout_carp'] = buyouts.get('Цап царапкин')
    metrics['buyout_hara'] = buyouts.get('Harakiri')

    success, report_id = save_report_to_db(
        file_name=f"auto_{period_str}.xlsx",
        file_hash=file_hash,
        date_period=period_str,
        start_date=date_from,
        end_date=date_to,
        values=values,
        metrics=metrics,
        articles={}
    )
    if not success:
        logger.error("Ошибка сохранения автоотчёта")
        return

    report_dir = Path("/data/reports")
    report_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, report_dir / f"отчёт_{report_id}.xlsx")

    def format_number(num):
        if num is None: return "0"
        if isinstance(num, float) and num.is_integer():
            return f"{int(num):,}".replace(",", " ")
        return f"{num:,.2f}".replace(",", " ")

    buyout_carp_str = f"{metrics['buyout_carp']:.1f}%" if metrics['buyout_carp'] is not None else "Н/Д"
    buyout_hara_str = f"{metrics['buyout_hara']:.1f}%" if metrics['buyout_hara'] is not None else "Н/Д"

    summary = (
        f"📊 Автоматический отчёт за {period_str}\n"
        f"💰 Оборот: {format_number(metrics.get('wb_total', 0))} ₽\n"
        f"🟢 ЦАП: {format_number(metrics.get('wb_carp', 0))} ₽\n"
        f"🔴 Harakiri: {format_number(metrics.get('wb_hara', 0))} ₽\n"
        f"💳 Эквайринг: {metrics.get('avg_acquiring', 0):.2f}%\n"
        f"💵 К выводу ЦАП: {format_number(metrics.get('k_vyvodu_carp', 0))} ₽\n"
        f"💵 К выводу Harakiri: {format_number(metrics.get('k_vyvodu_hara', 0))} ₽\n"
        f"📦 Выкуп ЦАП: {buyout_carp_str}\n"
        f"📦 Выкуп Harakiri: {buyout_hara_str}\n"
    )

    for uid in ALLOWED_USERS:
        try:
            with open(template_path, 'rb') as f:
                await app.bot.send_document(uid, document=f, filename=f"отчёт_{period_str}.xlsx",
                                           caption=f"✅ Автоматический отчёт за {period_str}")
            await app.bot.send_message(uid, summary, parse_mode='Markdown')
        except Exception as e:
            logger.error(f"Не удалось отправить автоотчёт пользователю {uid}: {e}")


async def fetch_reports_for_period(app, date_from, date_to, period_str, force=False):
    from wb_api import get_weekly_reports, get_report_detail

    if not force and get_report_id_by_period(date_from, date_to):
        msg = f"Отчёт за {period_str} уже существует. Используйте /wr с force для обновления."
        logger.info(msg)
        for uid in ALLOWED_USERS:
            try:
                await app.bot.send_message(uid, msg)
            except:
                pass
        return True

    reports_meta = get_weekly_reports(date_from, date_to)
    if not reports_meta:
        msg = f"Еженедельный отчёт за {period_str} ещё не готов."
        logger.info(msg)
        for uid in ALLOWED_USERS:
            try:
                await app.bot.send_message(uid, msg)
            except:
                pass
        return False

    osn_report = next((r for r in reports_meta if r['report_type'] == 1), None)
    vyk_report = next((r for r in reports_meta if r['report_type'] == 2), None)
    if not osn_report or not vyk_report:
        msg = f"Найдены не все типы отчётов за {period_str}."
        logger.warning(msg)
        for uid in ALLOWED_USERS:
            try:
                await app.bot.send_message(uid, msg)
            except:
                pass
        return False

    detail_osn = get_report_detail(osn_report['report_id'])
    detail_vyk = get_report_detail(vyk_report['report_id'])
    if not detail_osn or not detail_vyk:
        msg = f"Не удалось получить детализацию отчётов за {period_str}"
        logger.error(msg)
        for uid in ALLOWED_USERS:
            try:
                await app.bot.send_message(uid, msg)
            except:
                pass
        return False

    await process_auto_report(app, detail_osn, detail_vyk, period_str, date_from, date_to)
    return True


async def fetch_weekly_reports_job(app):
    today = datetime.now().date()
    last_monday = today - timedelta(days=today.weekday() + 7)
    last_sunday = last_monday + timedelta(days=6)
    date_from = last_monday.strftime("%Y-%m-%d")
    date_to = last_sunday.strftime("%Y-%m-%d")
    period_str = f"{last_monday.strftime('%d.%m')}-{last_sunday.strftime('%d.%m')}"

    logger.info(f"🔍 Автопроверка отчётов за {period_str}")
    success = await fetch_reports_for_period(app, date_from, date_to, period_str, force=False)

    if not success:
        scheduler.add_job(
            fetch_weekly_reports_job,
            'date',
            run_date=datetime.now() + timedelta(hours=1.5),
            args=[app],
            id=f"retry_{period_str}",
            replace_existing=True
        )
